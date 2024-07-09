import json
import time
import requests
from playwright.sync_api import sync_playwright
import os
import xml.etree.ElementTree as ET
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.hyperlink import Hyperlink

# 로그인 페이지 URL과 목표 페이지 URL
login_url = 'https://login.afreecatv.com/afreeca/login.php?szFrom=full&request_uri=https%3A%2F%2Fwww.afreecatv.com%2F'
url = input("url 주소 입력: ")

# 전역 변수 선언
all_timestamp = []
broadcast_title = ''
# 파일 경로 설정
desktop_path = os.path.expanduser("~/Desktop")
storage_file = os.path.join(desktop_path, 'code', 'getaf', 'storage_state.json')
result_folder = os.path.join(desktop_path, 'result')

# 결과 폴더가 존재하지 않으면 생성
if not os.path.exists(result_folder):
    os.makedirs(result_folder)


# 위험한 문자 리스트
dangerous_chars = ['=', '+', '-', '@', '\\']

def sanitize_string(value):
    if value is None:
        return ''
    if not isinstance(value, str):
        return f'{value}'
    if any(value.startswith(char) for char in dangerous_chars):
        return f' {value}'
    return f'{value}'


def format_timestamp(seconds):
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    return f"{round(hours):02}:{round(minutes):02}:{round(seconds):02}"

def save_login_state():
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        
        page.goto(login_url)
        
        print("로그인 후 계속 진행하려면 Enter 키를 누르세요...")
        input()
        
        context.storage_state(path=storage_file)
        browser.close()
        print("로그인 상태가 저장되었습니다.")

def extract_data():
    global broadcast_title
    try:
        with sync_playwright() as p:
            browser = p.firefox.launch(headless=True)
            context = browser.new_context(storage_state=storage_file)
            page = context.new_page()
            
            page.goto(url)
            time.sleep(2)

            title_selector = ".broadcast_title"
            title_element = page.query_selector(title_selector)
            if title_element:
                broadcast_title = title_element.get_attribute("title")
                print(f"Broadcast Title: {broadcast_title}")
            else:
                print("Broadcast title element not found.")

            script = """
            () => {
                function getCircularReplacer() {
                    const seen = new WeakSet();
                    return (key, value) => {
                        if (typeof value === "object" && value !== null) {
                            if (seen.has(value)) {
                                return;
                            }
                            seen.add(value);
                        }
                        return value;
                    };
                }
                return JSON.stringify(window.vodCore, getCircularReplacer());
            }
            """
            vodcore_data = page.evaluate(script)

            vodcore_json = json.loads(vodcore_data)

            row_keys = [item['fileInfoKey'] for item in vodcore_json['fileItems']]
            durations = [item['duration'] for item in vodcore_json['fileItems']]

            print(row_keys)
            print(durations)
            fetch_data(row_keys, durations)
            browser.close()
    except Exception as e:
        print(f"에러가 발생했습니다: {e}")
        save_login_state()
        extract_data()

url_template = 'https://videoimg.afreecatv.com/php/ChatLoadSplit.php?rowKey={row_key}_c&startTime={start_time}'

def fetch_data(row_keys, durations):
    file_counter = 1
    for row_key, duration in zip(row_keys, durations):
        iterations = duration // 300
        
        for start_time in range(0, (iterations + 1) * 300, 300):
            try:
                url = url_template.format(row_key=row_key, start_time=start_time)
                response = requests.get(url)
                response.raise_for_status()
                
                xml_data = response.content
                
                file_name = f"chat_{file_counter:04}.xml"
                file_counter += 1
                xml_file_path = os.path.join(result_folder, file_name)
                
                with open(xml_file_path, "wb") as xml_file:
                    xml_file.write(xml_data)

                print(f'{url}\n진행중: {xml_file_path}')
                time.sleep(0.7)
            except Exception as e:
                print(f"데이터 추출 중 오류가 발생했습니다: {str(e)}")
                continue

def extract_required_data_from_xml(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    extracted_data = []
    total_sum = 0
    
    for element in root:
        if element.tag in ['chat', 'balloon', 'adballoon', 'ogq']:
            nickname = ''
            user_id = ''
            message = ''
            timestamp = ''
            
            if element.tag == 'ogq':
                user_id = element.find('s').text if element.find('s') is not None else ''
                nickname = element.find('sn').text if element.find('sn') is not None else ''
                message = element.find('m').text if element.find('m') is not None else ''
                timestamp = element.find('t').text if element.find('t') is not None else ''
            elif element.tag in ['balloon', 'adballoon']:
                nickname = element.find('n').text if element.find('n') is not None else ''
                user_id = element.find('u').text if element.find('u') is not None else ''
                message = element.find('c').text if element.find('c') is not None else ''
                timestamp = element.find('t').text if element.find('t') is not None else ''
                
                if message and message.isdigit():
                    total_sum += int(message)
            else:
                nickname = element.find('n').text if element.find('n') is not None else ''
                user_id = element.find('u').text if element.find('u') is not None else ''
                message = element.find('m').text if element.find('m') is not None else ''
                timestamp = element.find('t').text if element.find('t') is not None else ''
            
            all_timestamp.append(timestamp)

            # 유저 아이디에서 괄호와 숫자 제거
            user_id = re.sub(r'\(\d+\)$', '', user_id)
            
            extracted_data.append({
                'tag': element.tag,
                'nickname': sanitize_string(nickname),
                'user_id': user_id,
                'message': sanitize_string(message),
                'timestamp': timestamp
            })
    
    return extracted_data, total_sum

def find_and_append_chat_messages(data, current_index):
    user_id = data[current_index]['user_id']
    chat_messages = []
    for i in range(current_index + 1, len(data)):
        if data[i]['tag'] == 'chat' and data[i]['user_id'] == user_id:
            chat_messages.append(data[i]['message'])
        if len(chat_messages) == 3:
            break
    return chat_messages

def process_all_xml_files():
    all_extracted_data = []
    total_sum = 0

    for filename in sorted(os.listdir(result_folder)):
        if filename.endswith('.xml'):
            file_path = os.path.join(result_folder, filename)
            extracted_data, file_sum = extract_required_data_from_xml(file_path)
            total_sum += file_sum
            all_extracted_data.extend(extracted_data)
    
    b = all_timestamp
    c = []

    prev_value = 0

    for i in range(len(b)):
        current_value = float(b[i])
        
        if i == 0:
            c.append(float(current_value))
        else:
            if current_value < float(b[i - 1]):
                prev_value = c[-1]
            c.append(float(prev_value + current_value))

    # Excel 파일 생성 부분
    wb = Workbook()
    ws = wb.active
    ws.title = "excel"

    # 헤더 추가 및 스타일 적용
    headers = ['Tag', 'Timestamp', 'Nickname', 'User ID', 'Balloons', 'Message', 'FollowMessage', 'Accumulated']
    header_fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill

    # 데이터 추가
    accumulated_sum = 0
    for idx, (data, timestamp) in enumerate(zip(all_extracted_data, c), start=2):
        message = data['message']
        balloon_value = 0
        follow_messages = []

        if data['tag'] in ['balloon', 'adballoon']:
            if message and message.isdigit():
                balloon_value = int(message)
                accumulated_sum += balloon_value

                if balloon_value >= 100:
                    message = ''
                    follow_messages = find_and_append_chat_messages(all_extracted_data, idx-2)
                    ws.cell(row=idx, column=5).fill = PatternFill(start_color="eaffe6", end_color="FFCCCB", fill_type="solid")
                    ws.cell(row=idx, column=6).fill = PatternFill(start_color="ffe9f3", end_color="FFCCCB", fill_type="solid")
                else:
                    message = ''
        # 각 셀에 데이터 추가 및 스타일 적용
        ws.cell(row=idx, column=1, value=data['tag']).alignment = Alignment(horizontal='center')
        timestamp_cell = ws.cell(row=idx, column=2, value=format_timestamp(timestamp))
        timestamp_cell.alignment = Alignment(horizontal='center')
        if balloon_value >= 300:
            timestamp_cell.hyperlink = Hyperlink(ref=f"B{idx}", target=f"{url}?change_second={round(timestamp)-3}")
            timestamp_cell.font = Font(color="6262ff", underline="single")
        ws.cell(row=idx, column=3, value=data['nickname']).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=4, value=data['user_id']).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=5, value=balloon_value if balloon_value >= 1 else 0).alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=6, value=follow_messages[0] if follow_messages else message)
        ws.cell(row=idx, column=7, value=" | ".join(follow_messages[1:3]) if len(follow_messages) > 1 else "")
        ws.cell(row=idx, column=8, value=accumulated_sum).alignment = Alignment(horizontal='center')

    # 열 너비 자동 조정 (최대 80으로 제한)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.2, 60)  # 최대 80으로 제한
        ws.column_dimensions[column_letter].width = adjusted_width

    # 필터링
    ws.auto_filter.ref = ws.dimensions
    # 파일 저장
    # excel_file_path = os.path.join(desktop_path, 'result', f'{broadcast_title.replace("/","")}.xlsx')
    # excel_file_path = os.path.join(desktop_path, 'result', f'{total_sum}개.xlsx')
    excel_file_path = os.path.join(desktop_path, 'result', '팔로우진행중58.xlsx')
    wb.save(excel_file_path)

    print(f"Excel 파일이 저장되었습니다: {excel_file_path}")
    print(f"Total Sum of balloon: {total_sum}")

# 메인 실행 부분
if not os.path.exists(storage_file):
    save_login_state()

extract_data()
process_all_xml_files()
