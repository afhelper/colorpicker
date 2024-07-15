import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment,PatternFill,Font
from openpyxl.worksheet.hyperlink import Hyperlink

# 사용자로부터 URL 입력받기
base_url = input("url 주소 입력: ")
# base_url = "1"
# 엑셀 파일 읽기
file_path = '/Users/air/Desktop/result/446개.xlsx'  # 원본 엑셀 파일 경로를 지정해주세요
# file_path = input("경로명 입력 :")
df = pd.read_excel(file_path)

# 필터링 조건 리스트
filter_conditions = [
    # (df['Message'].str.startswith('하') | df['Message'].str.endswith('하') | df['Message'].str.contains('트|뚱|ㅌㅇ|유진|하\+', na=False), '하'),
    (df['Message'].str.startswith('조') | df['Message'].str.endswith('조') | df['Message'].str.contains('조플|유정|ㅇㅈ', na=False), '조'),
    # (df['Message'].str.startswith('천') | df['Message'].str.endswith('천')  | (df['Message'].str == 'ㅊ') | df['Message'].str.contains('이주|주미|이죽|ㅇㅈ', na=False), '천'),
    # (df['Message'].str.startswith('쟌') | df['Message'].str.endswith('쟌') | df['Message'].str.contains('지안|ㅈㅇ|쟌', na=False), '쟌'),
    # (df['Message'].str.startswith('임') | df['Message'].str.contains('주연', na=False), '임'),
    (df['Message'].str.startswith('주성') | df['Message'].str.contains('주성|주플', na=False), '주')
    # (df['Message'].str.startswith('은') | df['Message'].str.endswith('은') | df['Message'].str.contains('유화|은플|응뉴', na=False), '은')
]

# 필터링 조건에 따라 파일을 생성하는 함수
def create_filtered_excel(filtered_df, label, base_url):
    balloons_sum = filtered_df['Balloons'].sum()
    output_file_path = f'/Users/air/Desktop/result/{label}_{balloons_sum}개.xlsx'
    filtered_df.to_excel(output_file_path, index=False)

    workbook = load_workbook(output_file_path)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):
        for cell in row:
            timestamp = cell.value
            if isinstance(timestamp, str):
                h, m, s = map(int, timestamp.split(':'))
                total_seconds = h * 3600 + m * 60 + s
                hyperlink_url = f"{base_url}?change_second={total_seconds}"
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=hyperlink_url)
                cell.style = "Hyperlink"

    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color="969696", end_color="969696", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    balloons_fill = PatternFill(start_color="eaffe6", end_color="eaffe6", fill_type="solid")
    message_fill = PatternFill(start_color="ffe9f3", end_color="ffe9f3", fill_type="solid")

    # 헤더 색상 채우기
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter not in ['F', 'G']:
                cell.alignment = center_alignment
            if cell.column_letter == 'E':  # 'Balloons' 열 색칠
                cell.fill = balloons_fill
            elif cell.column_letter == 'F':  # 'Message' 열 색칠
                cell.fill = message_fill

    for cell in worksheet[1]:
        cell.alignment = center_alignment

    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 5)
        column_letter = column[0].column_letter
        if column_letter == 'F' or column_letter == 'G':
            worksheet.column_dimensions[column_letter].width = 80
        elif column_letter == 'A' or column_letter == 'D' or column_letter == 'H':
            worksheet.column_dimensions[column_letter].hidden = True
        else:
            worksheet.column_dimensions[column_letter].width = adjusted_width

    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(output_file_path)

    print(f"필터링된 데이터를 '{output_file_path}'에 저장했습니다.")

# 모든 필터링 조건에 대해 처리
for condition, label in filter_conditions:
    filtered_df = df[(df['Balloons'] >= 100) & condition]
    create_filtered_excel(filtered_df, label, base_url)

# 필터링되지 않은 데이터 추가 처리
create_filtered_excel(df[df['Balloons'] >= 100], '100개이상', base_url)
create_filtered_excel(df[df['Balloons'] >= 1000], '1000개이상', base_url)
create_filtered_excel(df[df['Balloons'] >= 10000], '10000개이상', base_url)
